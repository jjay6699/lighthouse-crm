from __future__ import annotations

import glob
import os
import re
import sqlite3
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pypdf

ROOT = Path(__file__).resolve().parents[1]
DEBIT_NOTES_DIR = Path(os.environ.get("DEBIT_NOTES_DIR", ROOT / "debit notes"))
DATA_DIR = Path(os.environ.get("DATA_DIR", ROOT / "data"))
DB_PATH = Path(os.environ.get("DATABASE_PATH", DATA_DIR / "finance.sqlite"))


def parse_rate(val) -> float | None:
    if val is None:
        return None
    val_str = str(val).replace(",", "").strip()
    if not val_str:
        return None
    # Match the last float or integer in the string (handles transformations like $104 --> $71.5)
    match = re.search(r"(\d+(?:\.\d+)?)\s*$", val_str)
    if match:
        try:
            return float(match.group(1))
        except (TypeError, ValueError):
            return None
    return None


def parse_xlsm_date(val) -> str | None:
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, str):
        val = val.strip()
        # Handle formats like YYYYMMDD
        match_yyyymmdd = re.match(r"^(\d{4})(\d{2})(\d{2})$", val)
        if match_yyyymmdd:
            return f"{match_yyyymmdd.group(1)}-{match_yyyymmdd.group(2)}-{match_yyyymmdd.group(3)}"
        # Handle YYYY/MM/DD or YYYY-MM-DD
        match = re.match(r"^(\d{4})[/-](\d{2})[/-](\d{2})", val)
        if match:
            return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    return None


def parse_pdf_date(val_str: str) -> str | None:
    if not val_str or len(val_str) != 8:
        return None
    try:
        return f"{val_str[0:4]}-{val_str[4:6]}-{val_str[6:8]}"
    except Exception:
        return None


def init_debit_note_tables(conn: sqlite3.Connection):
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS debit_notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT NOT NULL,
            sku TEXT NOT NULL,
            description TEXT NOT NULL,
            qty INTEGER NOT NULL,
            unit_cost REAL NOT NULL,
            date_from TEXT NOT NULL,
            date_to TEXT NOT NULL,
            raw_line_1 TEXT,
            raw_line_2 TEXT
        );

        CREATE TABLE IF NOT EXISTS promotion_proposals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT NOT NULL,
            sku TEXT NOT NULL,
            brand TEXT NOT NULL,
            description TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            rsp REAL,
            promo_price REAL,
            mix_match_offer TEXT,
            final_avg_sp REAL,
            invoice_unit_cost REAL,
            funding_support_pc REAL NOT NULL,
            adjustment_basis TEXT,
            funding_type TEXT
        );

        CREATE INDEX IF NOT EXISTS idx_debit_notes_sku ON debit_notes(sku);
        CREATE INDEX IF NOT EXISTS idx_debit_notes_dates ON debit_notes(date_from, date_to);
        CREATE INDEX IF NOT EXISTS idx_promotion_proposals_sku ON promotion_proposals(sku);
        CREATE INDEX IF NOT EXISTS idx_promotion_proposals_dates ON promotion_proposals(start_date, end_date);
        """
    )
    # Clear existing entries for fresh idempotent run
    conn.execute("DELETE FROM debit_notes")
    conn.execute("DELETE FROM promotion_proposals")
    conn.commit()


def import_pdfs(conn: sqlite3.Connection):
    pdf_files = glob.glob(str(DEBIT_NOTES_DIR / "*.pdf"))
    print(f"Found {len(pdf_files)} PDF debit note files to process.")
    
    total_imported = 0
    
    for fp in pdf_files:
        file_name = os.path.basename(fp)
        try:
            reader = pypdf.PdfReader(fp)
            full_text = ""
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"
            
            lines = full_text.split("\n")
            
            # Match quantity & unit cost, e.g. "243 pcs X $71.50" or "1 pcs X $73.30"
            qty_price_pattern = re.compile(r"(\d+)\s+pcs?\s+X\s+\$(\d+\.\d{2})")
            # Match sku & date range, e.g. "|I821219|DF20260401|DT20260409|"
            meta_pattern = re.compile(r"\|I?(\d+)\|DF(\d{8})\|DT(\d{8})")
            
            for idx, line in enumerate(lines):
                qty_price_match = qty_price_pattern.search(line)
                if qty_price_match:
                    # Check the next line for meta details
                    if idx + 1 < len(lines):
                        next_line = lines[idx+1]
                        meta_match = meta_pattern.search(next_line)
                        if meta_match:
                            sku = meta_match.group(1)
                            date_from = parse_pdf_date(meta_match.group(2))
                            date_to = parse_pdf_date(meta_match.group(3))
                            qty = int(qty_price_match.group(1))
                            unit_cost = float(qty_price_match.group(2))
                            
                            # Clean product description from the first line
                            # Usually of style: "OT HEALTH               森暉去濕消水丸 60粒裝 52 pcs X $71.50 for"
                            desc_clean = line
                            # Remove prefix like "OT HEALTH", "OT SKIN & COSMETIC", etc.
                            desc_clean = re.sub(r"^OT\s+[A-Za-z\s&]+\s+", "", desc_clean)
                            # Remove suffix with quantity and price
                            desc_clean = qty_price_pattern.sub("", desc_clean).replace("for", "").strip()
                            
                            conn.execute(
                                """
                                INSERT INTO debit_notes (
                                    file_name, sku, description, qty, unit_cost, date_from, date_to, raw_line_1, raw_line_2
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                (file_name, sku, desc_clean, qty, unit_cost, date_from, date_to, line.strip(), next_line.strip())
                            )
                            total_imported += 1
            print(f"Successfully processed PDF: {file_name}")
        except Exception as e:
            print(f"Error processing PDF {file_name}: {e}")
            
    conn.commit()
    print(f"Imported {total_imported} line items from PDF files.")


def import_xlsms(conn: sqlite3.Connection):
    xlsm_files = glob.glob(str(DEBIT_NOTES_DIR / "*.xlsm"))
    print(f"Found {len(xlsm_files)} XLSM promotion proposal files to process.")
    
    total_imported = 0
    
    for fp in xlsm_files:
        file_name = os.path.basename(fp)
        try:
            wb = openpyxl.load_workbook(fp, data_only=True)
            if 'FORM HK&Macau' not in wb.sheetnames:
                print(f"Skipping {file_name}: 'FORM HK&Macau' sheet not found.")
                continue
                
            sheet = wb['FORM HK&Macau']
            
            # Row 11 has headers. Let's make sure it contains code and dates
            row_11_vals = [sheet.cell(11, c).value for c in range(1, sheet.max_column + 1)]
            
            # Find column indexes (1-based)
            col_sku = None
            col_brand = None
            col_desc = None
            col_start = None
            col_end = None
            col_rsp = None
            col_promo_price = None
            col_mix_match = None
            col_final_sp = None
            col_unit_cost = None
            col_support = None
            col_adjust = None
            col_type = None
            
            for idx, h in enumerate(row_11_vals, 1):
                if not h:
                    continue
                h_str = str(h).lower()
                if "code" in h_str and col_sku is None:
                    col_sku = idx
                elif "brand" in h_str and col_brand is None:
                    col_brand = idx
                elif "description" in h_str and col_desc is None:
                    col_desc = idx
                elif "start date" in h_str and col_start is None:
                    # Capture first start date
                    col_start = idx
                elif "end date" in h_str and col_end is None:
                    # Capture first end date
                    col_end = idx
                elif "recommended selling" in h_str or "rsp" in h_str:
                    col_rsp = idx
                elif "promotion price" in h_str:
                    col_promo_price = idx
                elif "mix & match" in h_str or "mechanics offer" in h_str:
                    col_mix_match = idx
                elif "final average selling" in h_str:
                    col_final_sp = idx
                elif "invoice unit cost" in h_str:
                    col_unit_cost = idx
                elif "funding support per pc" in h_str and "(setting" not in h_str:
                    col_support = idx
                elif "adjustment basis" in h_str:
                    col_adjust = idx
                elif "funding support per pc" in h_str and "setting" in h_str:
                    col_type = idx

            # Fallback indexes based on our standard discovered layout if any are missing
            if col_sku is None: col_sku = 2
            if col_brand is None: col_brand = 3
            if col_desc is None: col_desc = 4
            if col_start is None: col_start = 6
            if col_end is None: col_end = 7
            if col_rsp is None: col_rsp = 8
            if col_promo_price is None: col_promo_price = 9
            if col_mix_match is None: col_mix_match = 10
            if col_final_sp is None: col_final_sp = 11
            if col_unit_cost is None: col_unit_cost = 21
            if col_support is None: col_support = 25
            if col_type is None: col_type = 29
            if col_adjust is None: col_adjust = 31
            
            # Read rows starting from row 12
            for r in range(12, sheet.max_row + 1):
                sku_raw = sheet.cell(r, col_sku).value
                if sku_raw is None:
                    continue
                    
                # Clean SKU to ensure it's a solid code
                sku = str(sku_raw).strip()
                # Split decimals if openpyxl read as float e.g. 821219.0
                if sku.endswith(".0"):
                    sku = sku[:-2]
                    
                if not sku or not sku.isdigit():
                    continue
                
                # Check for support rate
                support_raw = sheet.cell(r, col_support).value
                support_rate = parse_rate(support_raw)
                if support_rate is None:
                    # Skip proposals with no support rate (not a trade promotion charge or not active)
                    continue
                
                brand = str(sheet.cell(r, col_brand).value or "").strip()
                desc = str(sheet.cell(r, col_desc).value or "").strip()
                
                start_date = parse_xlsm_date(sheet.cell(r, col_start).value)
                end_date = parse_xlsm_date(sheet.cell(r, col_end).value)
                
                if not start_date or not end_date:
                    continue
                
                rsp = parse_rate(sheet.cell(r, col_rsp).value)
                promo_price = parse_rate(sheet.cell(r, col_promo_price).value)
                mix_match = str(sheet.cell(r, col_mix_match).value or "").strip()
                final_sp = parse_rate(sheet.cell(r, col_final_sp).value)
                unit_cost = parse_rate(sheet.cell(r, col_unit_cost).value)
                adjust = str(sheet.cell(r, col_adjust).value or "").strip()
                fund_type = str(sheet.cell(r, col_type).value or "").strip()
                
                conn.execute(
                    """
                    INSERT INTO promotion_proposals (
                        file_name, sku, brand, description, start_date, end_date, rsp, promo_price,
                        mix_match_offer, final_avg_sp, invoice_unit_cost, funding_support_pc,
                        adjustment_basis, funding_type
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        file_name, sku, brand, desc, start_date, end_date, rsp, promo_price,
                        mix_match, final_sp, unit_cost, support_rate, adjust, fund_type
                    )
                )
                total_imported += 1
            print(f"Successfully processed XLSM: {file_name}")
        except Exception as e:
            print(f"Error processing XLSM {file_name}: {e}")
            
    conn.commit()
    print(f"Imported {total_imported} proposal rows from XLSM files.")


def main():
    print("Starting Debit Note data import...")
    print(f"Database file: {DB_PATH}")
    print(f"Debit notes folder: {DEBIT_NOTES_DIR}")
    
    # Ensure data directory exists
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    
    conn = sqlite3.connect(str(DB_PATH))
    try:
        init_debit_note_tables(conn)
        import_pdfs(conn)
        import_xlsms(conn)
        print("Debit Note data import finished successfully!")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
